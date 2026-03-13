"""
export.py
---------
Exportación de barras a Excel en formato aSa Studio.

Genera un archivo ZIP con un .xlsx por cada combinación SECTOR+PISO+CICLO.
Cada archivo lleva el nombre: "{SECTOR} {PISO} {CICLO}.xlsx"

Endpoint:
- GET /proyectos/{id_proyecto}/exportar
"""

from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from io import BytesIO
import zipfile

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .db import get_conn, audit
from .auth import get_current_user

router = APIRouter()

# Columnas del formato de exportación aSa Studio
EXPORT_COLUMNS = [
    ("EJE",       "eje",          "text"),
    ("SECTOR",    "sector",       "text"),
    ("PISO",      "piso",         "text"),
    ("CICLO",     "ciclo",        "text"),
    ("CANT",      "cant_total",   "int"),
    ("Ømm",       "diam",         "int"),
    ("FIGURA",    "figura",       "text"),
    ("L/cm",      "largo_total",  "int"),
    ("MARCA",     "marca",        "text"),
    ("PROD",      "cod_proyecto", "text"),
    ("A cm",      "dim_a",        "num"),
    ("B cm",      "dim_b",        "num"),
    ("C cm",      "dim_c",        "num"),
    ("D cm",      "dim_d",        "num"),
    ("E cm",      "dim_e",        "num"),
    ("F cm",      "dim_f",        "num"),
    ("G cm",      "dim_g",        "num"),
    ("H cm",      "dim_h",        "num"),
    ("I cm",      "dim_i",        "num"),
    ("J cm",      None,           "num"),   # No existe en CSV, siempre vacío
    ("AngV",      "ang1",         "ang"),
    ("AngV1",     "ang2",         "ang"),
    ("AngV2",     "ang3",         "ang"),
    ("AngV3",     "ang4",         "ang"),
    ("R cm",      "radio",        "num"),
    ("PesoKg",    "peso_unitario","dec3"),
    ("PesoTotal", "peso_total",   "dec2"),
]

# Columnas a consultar de la BD
DB_FIELDS = [
    "eje", "sector", "piso", "ciclo", "cant_total", "diam", "figura",
    "largo_total", "marca", "cod_proyecto",
    "dim_a", "dim_b", "dim_c", "dim_d", "dim_e", "dim_f", "dim_g", "dim_h", "dim_i",
    "ang1", "ang2", "ang3", "ang4", "radio",
    "peso_unitario", "peso_total",
]


def _build_sheet(wb: Workbook, sheet_name: str, rows: list, sector: str, piso: str, ciclo: str):
    """Construye una hoja Excel con el formato aSa Studio."""
    ws = wb.create_sheet(title=sheet_name[:31])  # Excel limita a 31 chars

    # Filas 1-4 vacías (espacio para metadata si se necesita)
    for i in range(1, 5):
        ws.cell(row=i, column=1, value="")

    # Fila 5: headers
    header_font = Font(bold=True, size=10)
    thin_border = Border(
        bottom=Side(style='thin')
    )
    for col_idx, (label, _, _) in enumerate(EXPORT_COLUMNS, 1):
        cell = ws.cell(row=5, column=col_idx, value=label)
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # Filas de datos a partir de fila 6
    for row_idx, row_data in enumerate(rows, 6):
        for col_idx, (_, db_col, fmt) in enumerate(EXPORT_COLUMNS, 1):
            if db_col is None:
                val = None
            else:
                val = row_data.get(db_col)

            if val is None:
                ws.cell(row=row_idx, column=col_idx, value=0 if fmt in ("int", "num", "dec1", "dec2", "ang") else "")
                continue

            if fmt == "int":
                try:
                    ws.cell(row=row_idx, column=col_idx, value=int(round(float(val))))
                except (ValueError, TypeError):
                    ws.cell(row=row_idx, column=col_idx, value=0)
            elif fmt == "dec1":
                try:
                    cell = ws.cell(row=row_idx, column=col_idx, value=round(float(val), 1))
                    cell.number_format = '0.0'
                except (ValueError, TypeError):
                    ws.cell(row=row_idx, column=col_idx, value=0)
            elif fmt == "dec2":
                try:
                    cell = ws.cell(row=row_idx, column=col_idx, value=round(float(val), 2))
                    cell.number_format = '0.00'
                except (ValueError, TypeError):
                    ws.cell(row=row_idx, column=col_idx, value=0)
            elif fmt == "dec3":
                try:
                    cell = ws.cell(row=row_idx, column=col_idx, value=round(float(val), 3))
                    cell.number_format = '0.000'
                except (ValueError, TypeError):
                    ws.cell(row=row_idx, column=col_idx, value=0)
            elif fmt == "num":
                try:
                    ws.cell(row=row_idx, column=col_idx, value=float(val))
                except (ValueError, TypeError):
                    ws.cell(row=row_idx, column=col_idx, value=0)
            elif fmt == "ang":
                try:
                    fval = float(val)
                    ws.cell(row=row_idx, column=col_idx, value=f"<{int(fval)}" if fval != 0 else 0)
                except (ValueError, TypeError):
                    ws.cell(row=row_idx, column=col_idx, value=0)
            else:
                ws.cell(row=row_idx, column=col_idx, value=str(val) if val else "")

    # Ajustar ancho de columnas
    for col_idx, (label, _, _) in enumerate(EXPORT_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(label) + 2, 8)


@router.get("/proyectos/{id_proyecto}/exportar")
def exportar_proyecto(
    id_proyecto: str,
    sectores: Optional[str] = Query(None, description="Comma-separated SECTOR_PISO_CICLO keys to export selectively"),
    user=Depends(get_current_user),
):
    """Exportar barras de un proyecto como ZIP de archivos Excel (uno por SECTOR+PISO+CICLO).
    
    Si `sectores` se proporciona (ej: 'ELEV_P1_C1,FUND_P1_C2'), solo se exportan
    esas combinaciones. Si está vacío o ausente, se exporta todo.
    """

    # Parse selective filter
    selected_set = None
    if sectores and sectores.strip():
        selected_set = set()
        for key in sectores.split(','):
            parts = key.strip().split('_', 2)  # SECTOR_PISO_CICLO
            if len(parts) == 3:
                selected_set.add(tuple(parts))

    fields_sql = ", ".join(DB_FIELDS)

    with get_conn() as conn:
        with conn.cursor() as cur:
            # Verificar que el proyecto existe
            cur.execute("SELECT nombre_proyecto FROM proyectos WHERE id_proyecto = %s", (id_proyecto,))
            proy = cur.fetchone()
            if not proy:
                raise HTTPException(status_code=404, detail="Proyecto no encontrado")
            nombre_proyecto = proy[0]

            # Obtener combinaciones únicas de sector+piso+ciclo
            cur.execute("""
                SELECT DISTINCT sector, piso, ciclo
                FROM barras
                WHERE id_proyecto = %s AND sector IS NOT NULL AND sector != ''
                ORDER BY sector, piso, ciclo
            """, (id_proyecto,))
            combos = cur.fetchall()

            # Filter by selected combinations if provided
            if selected_set is not None:
                combos = [c for c in combos if (c[0], c[1], c[2]) in selected_set]

            if not combos:
                raise HTTPException(status_code=404, detail="No hay barras para exportar en este proyecto")

            # Construir ZIP en memoria
            zip_buffer = BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
                for sector, piso, ciclo in combos:
                    cur.execute(f"""
                        SELECT {fields_sql}
                        FROM barras
                        WHERE id_proyecto = %s AND sector = %s AND piso = %s AND ciclo = %s
                        ORDER BY eje, diam, cant_total
                    """, (id_proyecto, sector, piso, ciclo))

                    col_names = [desc[0] for desc in cur.description]
                    rows = [dict(zip(col_names, r)) for r in cur.fetchall()]

                    if not rows:
                        continue

                    # Crear workbook para esta combinación
                    wb = Workbook()
                    # Eliminar hoja por defecto
                    wb.remove(wb.active)

                    sheet_name = f"{sector} {piso} {ciclo}"
                    _build_sheet(wb, sheet_name, rows, sector, piso, ciclo)

                    # Guardar workbook en buffer
                    xlsx_buffer = BytesIO()
                    wb.save(xlsx_buffer)
                    xlsx_buffer.seek(0)

                    filename = f"{sector} {piso} {ciclo}.xlsx"
                    zf.writestr(filename, xlsx_buffer.getvalue())

            zip_buffer.seek(0)

            # Log each exported sector to export_log
            email = user.get("email", "unknown")
            now_iso = __import__("datetime").datetime.utcnow().isoformat() + "Z"
            for sector, piso, ciclo in combos:
                export_key = f"{sector}_{piso}_{ciclo}"
                # Count barras/kilos for this combo
                cur.execute("""
                    SELECT COUNT(*), COALESCE(SUM(peso_total), 0)
                    FROM barras WHERE id_proyecto = %s AND sector = %s AND piso = %s AND ciclo = %s
                """, (id_proyecto, sector, piso, ciclo))
                stats = cur.fetchone()
                cur.execute("""
                    INSERT INTO export_log (id_proyecto, sector, piso, ciclo, export_key, usuario, fecha, barras, kilos)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (id_proyecto, sector, piso, ciclo, export_key, email, now_iso,
                      int(stats[0]) if stats else 0, round(float(stats[1]), 2) if stats else 0))

            audit(email, "exportar_excel", f"{len(combos)} sectores, {nombre_proyecto}", "proyecto", id_proyecto)

            # Nombre del ZIP: nombre del proyecto (sanitizado)
            safe_name = "".join(c for c in nombre_proyecto if c.isalnum() or c in " -_").strip()
            zip_filename = f"{safe_name or id_proyecto}_export.zip"

            return StreamingResponse(
                zip_buffer,
                media_type="application/zip",
                headers={"Content-Disposition": f'attachment; filename="{zip_filename}"'}
            )


@router.get("/proyectos/{id_proyecto}/export-history")
def export_history(
    id_proyecto: str,
    user=Depends(get_current_user),
):
    """Retorna historial de exportaciones por sector para un proyecto.
    Para cada export_key, indica si fue exportado, cuántas veces y la última fecha.
    También incluye ultima_modificacion para detectar sectores modificados post-exportación."""
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT 1 FROM proyectos WHERE id_proyecto = %s", (id_proyecto,))
            if not cur.fetchone():
                raise HTTPException(status_code=404, detail="Proyecto no encontrado")

            # Export log history
            cur.execute("""
                SELECT export_key,
                       COUNT(*) AS veces,
                       MAX(fecha) AS ultima_fecha,
                       (array_agg(usuario ORDER BY fecha DESC))[1] AS ultimo_usuario,
                       SUM(barras) AS total_barras,
                       SUM(kilos) AS total_kilos
                FROM export_log
                WHERE id_proyecto = %s
                GROUP BY export_key
                ORDER BY export_key
            """, (id_proyecto,))
            rows = cur.fetchall()
            
            # Get last modification date per sector (max fecha_carga)
            cur.execute("""
                SELECT UPPER(sector) || '_' || piso || '_' || ciclo AS export_key,
                       MAX(fecha_carga) AS ultima_modificacion,
                       COUNT(*) AS barras_actuales
                FROM barras
                WHERE id_proyecto = %s AND sector IS NOT NULL AND sector != ''
                GROUP BY UPPER(sector), piso, ciclo
            """, (id_proyecto,))
            mod_rows = cur.fetchall()
            mod_map = {r[0]: {"ultima_modificacion": r[1], "barras_actuales": int(r[2])} for r in mod_rows}

    history = {}
    for r in rows:
        key = r[0]
        mod_info = mod_map.get(key, {})
        history[key] = {
            "veces": int(r[1]),
            "ultima_fecha": r[2],
            "ultimo_usuario": r[3],
            "total_barras": int(r[4] or 0),
            "total_kilos": round(float(r[5] or 0), 2),
            "ultima_modificacion": mod_info.get("ultima_modificacion"),
            "barras_actuales": mod_info.get("barras_actuales", 0),
        }
    
    return {
        "id_proyecto": id_proyecto,
        "history": history
    }


@router.get("/proyectos/{id_proyecto}/export-report")
def export_report(
    id_proyecto: str,
    user=Depends(get_current_user),
):
    """Reporte completo: todos los sectores del proyecto con estado de exportación."""
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT nombre_proyecto FROM proyectos WHERE id_proyecto = %s", (id_proyecto,))
            prow = cur.fetchone()
            if not prow:
                raise HTTPException(status_code=404, detail="Proyecto no encontrado")

            # All sector combos in project
            cur.execute("""
                SELECT sector, piso, ciclo, COUNT(*) AS barras, COALESCE(SUM(peso_total), 0) AS kilos
                FROM barras
                WHERE id_proyecto = %s AND sector IS NOT NULL AND sector != ''
                GROUP BY sector, piso, ciclo
                ORDER BY sector, piso, ciclo
            """, (id_proyecto,))
            combos = cur.fetchall()

            # Export history
            cur.execute("""
                SELECT export_key, COUNT(*), MAX(fecha),
                       (array_agg(usuario ORDER BY fecha DESC))[1]
                FROM export_log WHERE id_proyecto = %s
                GROUP BY export_key
            """, (id_proyecto,))
            history = {r[0]: {"veces": int(r[1]), "ultima_fecha": r[2], "ultimo_usuario": r[3]} for r in cur.fetchall()}

    items = []
    for sector, piso, ciclo, barras, kilos in combos:
        key = f"{sector}_{piso}_{ciclo}"
        h = history.get(key)
        items.append({
            "sector": sector, "piso": piso, "ciclo": ciclo,
            "export_key": key,
            "barras": int(barras), "kilos": round(float(kilos), 2),
            "exportado": h is not None,
            "veces_exportado": h["veces"] if h else 0,
            "ultima_fecha": h["ultima_fecha"] if h else None,
            "ultimo_usuario": h["ultimo_usuario"] if h else None,
        })

    total = len(items)
    exportados = sum(1 for i in items if i["exportado"])

    return {
        "id_proyecto": id_proyecto,
        "nombre_proyecto": prow[0],
        "total_sectores": total,
        "exportados": exportados,
        "pendientes": total - exportados,
        "porcentaje": round(exportados / total * 100, 1) if total > 0 else 0,
        "items": items,
    }
